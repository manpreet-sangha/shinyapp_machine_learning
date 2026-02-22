"""
Add official trading holiday sheets to IDX_data.xlsx for each market.
Covers the data range: 2024-02-14 to 2026-02-20

Markets:
  - USA_Holidays      : NYSE/NASDAQ (INDU, SPX, SP_VIX)
  - Germany_Holidays   : XETRA (DAX)
  - France_Holidays    : Euronext Paris (CAC)
  - Japan_Holidays     : TSE (NKY)
  - HongKong_Holidays  : HKEX (HSI)
  - India_Holidays     : NSE (NIFTY, INVIXN) — already exists as NIFTY_Holidays
"""

import pandas as pd
import os
from openpyxl import load_workbook

input_fp = r"C:\Users\Manpreet\OneDrive - City St George's, University of London\Documents\Term 2\SMM748 Machine Learning Quantitative Finance\First Group Coursework ML\MS_Python_CW\input_data"
file_path = os.path.join(input_fp, 'IDX_data.xlsx')

# ──────────────────────────────────────────────────────────
# NYSE / NASDAQ — USA Holidays (market closed)
# ──────────────────────────────────────────────────────────
usa_holidays = [
    # 2024
    ("2024-02-19", "Mon", "Presidents' Day"),
    ("2024-03-29", "Fri", "Good Friday"),
    ("2024-05-27", "Mon", "Memorial Day"),
    ("2024-06-19", "Wed", "Juneteenth"),
    ("2024-07-04", "Thu", "Independence Day"),
    ("2024-09-02", "Mon", "Labor Day"),
    ("2024-11-28", "Thu", "Thanksgiving Day"),
    ("2024-12-25", "Wed", "Christmas Day"),
    # 2025
    ("2025-01-01", "Wed", "New Year's Day"),
    ("2025-01-09", "Thu", "National Day of Mourning (Jimmy Carter)"),
    ("2025-01-20", "Mon", "Martin Luther King Jr. Day"),
    ("2025-02-17", "Mon", "Presidents' Day"),
    ("2025-04-18", "Fri", "Good Friday"),
    ("2025-05-26", "Mon", "Memorial Day"),
    ("2025-06-19", "Thu", "Juneteenth"),
    ("2025-07-04", "Fri", "Independence Day"),
    ("2025-09-01", "Mon", "Labor Day"),
    ("2025-11-27", "Thu", "Thanksgiving Day"),
    ("2025-12-25", "Thu", "Christmas Day"),
    # 2026
    ("2026-01-01", "Thu", "New Year's Day"),
    ("2026-01-19", "Mon", "Martin Luther King Jr. Day"),
    ("2026-02-16", "Mon", "Presidents' Day"),
]

# ──────────────────────────────────────────────────────────
# XETRA — Germany Holidays (market closed)
# ──────────────────────────────────────────────────────────
germany_holidays = [
    # 2024
    ("2024-03-29", "Fri", "Good Friday"),
    ("2024-04-01", "Mon", "Easter Monday"),
    ("2024-05-01", "Wed", "Labour Day"),
    ("2024-12-24", "Tue", "Christmas Eve"),
    ("2024-12-25", "Wed", "Christmas Day"),
    ("2024-12-26", "Thu", "St. Stephen's Day"),
    ("2024-12-31", "Tue", "New Year's Eve"),
    # 2025
    ("2025-01-01", "Wed", "New Year's Day"),
    ("2025-04-18", "Fri", "Good Friday"),
    ("2025-04-21", "Mon", "Easter Monday"),
    ("2025-05-01", "Thu", "Labour Day"),
    ("2025-12-24", "Wed", "Christmas Eve"),
    ("2025-12-25", "Thu", "Christmas Day"),
    ("2025-12-26", "Fri", "St. Stephen's Day"),
    ("2025-12-31", "Wed", "New Year's Eve"),
    # 2026
    ("2026-01-01", "Thu", "New Year's Day"),
]

# ──────────────────────────────────────────────────────────
# Euronext Paris — France Holidays (market closed)
# ──────────────────────────────────────────────────────────
france_holidays = [
    # 2024
    ("2024-03-29", "Fri", "Good Friday"),
    ("2024-04-01", "Mon", "Easter Monday"),
    ("2024-05-01", "Wed", "Labour Day"),
    ("2024-12-24", "Tue", "Christmas Eve"),
    ("2024-12-25", "Wed", "Christmas Day"),
    ("2024-12-26", "Thu", "St. Stephen's Day"),
    ("2024-12-31", "Tue", "New Year's Eve"),
    # 2025
    ("2025-01-01", "Wed", "New Year's Day"),
    ("2025-04-18", "Fri", "Good Friday"),
    ("2025-04-21", "Mon", "Easter Monday"),
    ("2025-05-01", "Thu", "Labour Day"),
    ("2025-12-24", "Wed", "Christmas Eve"),
    ("2025-12-25", "Thu", "Christmas Day"),
    ("2025-12-26", "Fri", "St. Stephen's Day"),
    ("2025-12-31", "Wed", "New Year's Eve"),
    # 2026
    ("2026-01-01", "Thu", "New Year's Day"),
]

# ──────────────────────────────────────────────────────────
# TSE — Japan Holidays (market closed)
# ──────────────────────────────────────────────────────────
japan_holidays = [
    # 2024
    ("2024-02-23", "Fri", "Emperor's Birthday"),
    ("2024-03-20", "Wed", "Vernal Equinox Day"),
    ("2024-04-29", "Mon", "Showa Day"),
    ("2024-05-03", "Fri", "Constitution Memorial Day"),
    ("2024-05-06", "Mon", "Children's Day (observed)"),
    ("2024-07-15", "Mon", "Marine Day"),
    ("2024-08-12", "Mon", "Mountain Day (observed)"),
    ("2024-09-16", "Mon", "Respect for the Aged Day"),
    ("2024-09-23", "Mon", "Autumnal Equinox Day (observed)"),
    ("2024-10-14", "Mon", "Sports Day"),
    ("2024-11-04", "Mon", "Culture Day (observed)"),
    ("2024-12-31", "Tue", "New Year's Eve (market closed)"),
    # 2025
    ("2025-01-01", "Wed", "New Year's Day"),
    ("2025-01-02", "Thu", "New Year Bank Holiday"),
    ("2025-01-03", "Fri", "New Year Bank Holiday"),
    ("2025-01-13", "Mon", "Coming of Age Day"),
    ("2025-02-11", "Tue", "National Foundation Day"),
    ("2025-02-24", "Mon", "Emperor's Birthday (observed)"),
    ("2025-03-20", "Thu", "Vernal Equinox Day"),
    ("2025-04-29", "Tue", "Showa Day"),
    ("2025-05-03", "Sat", "Constitution Memorial Day"),
    ("2025-05-05", "Mon", "Children's Day"),
    ("2025-05-06", "Tue", "Greenery Day (observed)"),
    ("2025-07-21", "Mon", "Marine Day"),
    ("2025-08-11", "Mon", "Mountain Day"),
    ("2025-09-15", "Mon", "Respect for the Aged Day"),
    ("2025-09-23", "Tue", "Autumnal Equinox Day"),
    ("2025-10-13", "Mon", "Sports Day"),
    ("2025-11-03", "Mon", "Culture Day"),
    ("2025-11-24", "Mon", "Labour Thanksgiving Day (observed)"),
    ("2025-12-31", "Wed", "New Year's Eve (market closed)"),
    # 2026
    ("2026-01-01", "Thu", "New Year's Day"),
    ("2026-01-02", "Fri", "New Year Bank Holiday"),
    ("2026-01-12", "Mon", "Coming of Age Day"),
    ("2026-02-11", "Wed", "National Foundation Day"),
]

# ──────────────────────────────────────────────────────────
# HKEX — Hong Kong Holidays (market closed)
# ──────────────────────────────────────────────────────────
hongkong_holidays = [
    # 2024
    ("2024-02-12", "Mon", "Lunar New Year (Day 3)"),
    ("2024-02-13", "Tue", "Lunar New Year (Day 4 - extra)"),
    ("2024-03-29", "Fri", "Good Friday"),
    ("2024-04-01", "Mon", "Easter Monday"),
    ("2024-04-04", "Thu", "Ching Ming Festival"),
    ("2024-05-01", "Wed", "Labour Day"),
    ("2024-05-15", "Wed", "Buddha's Birthday"),
    ("2024-06-10", "Mon", "Tuen Ng Festival"),
    ("2024-07-01", "Mon", "HKSAR Establishment Day"),
    ("2024-09-18", "Wed", "Day after Mid-Autumn Festival"),
    ("2024-10-01", "Tue", "National Day"),
    ("2024-10-11", "Fri", "Chung Yeung Festival"),
    ("2024-12-25", "Wed", "Christmas Day"),
    ("2024-12-26", "Thu", "Boxing Day"),
    # 2025
    ("2025-01-01", "Wed", "New Year's Day"),
    ("2025-01-29", "Wed", "Lunar New Year (Day 1)"),
    ("2025-01-30", "Thu", "Lunar New Year (Day 2)"),
    ("2025-01-31", "Fri", "Lunar New Year (Day 3)"),
    ("2025-04-04", "Fri", "Ching Ming Festival"),
    ("2025-04-18", "Fri", "Good Friday"),
    ("2025-04-21", "Mon", "Easter Monday"),
    ("2025-05-01", "Thu", "Labour Day"),
    ("2025-05-05", "Mon", "Buddha's Birthday"),
    ("2025-05-31", "Sat", "Tuen Ng Festival"),
    ("2025-07-01", "Tue", "HKSAR Establishment Day"),
    ("2025-10-01", "Wed", "National Day"),
    ("2025-10-07", "Tue", "Day after Mid-Autumn Festival"),
    ("2025-10-29", "Wed", "Chung Yeung Festival"),
    ("2025-12-25", "Thu", "Christmas Day"),
    ("2025-12-26", "Fri", "Boxing Day"),
    # 2026
    ("2026-01-01", "Thu", "New Year's Day"),
    ("2026-02-17", "Tue", "Lunar New Year (Day 1)"),
    ("2026-02-18", "Wed", "Lunar New Year (Day 2)"),
    ("2026-02-19", "Thu", "Lunar New Year (Day 3)"),
]

# ──────────────────────────────────────────────────────────
# Build DataFrames and write to Excel
# ──────────────────────────────────────────────────────────
def make_holiday_df(holidays_list):
    df = pd.DataFrame(holidays_list, columns=['Date', 'Day', 'Holiday'])
    df['Date'] = pd.to_datetime(df['Date'])
    return df

sheets_to_add = {
    'USA_Holidays': make_holiday_df(usa_holidays),
    'Germany_Holidays': make_holiday_df(germany_holidays),
    'France_Holidays': make_holiday_df(france_holidays),
    'Japan_Holidays': make_holiday_df(japan_holidays),
    'HongKong_Holidays': make_holiday_df(hongkong_holidays),
}

# Load existing workbook and append new sheets
wb = load_workbook(file_path)

for sheet_name, df in sheets_to_add.items():
    # Remove sheet if it already exists (to avoid duplicates on re-run)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    # Write header
    ws.append(list(df.columns))
    # Write data
    for _, row in df.iterrows():
        ws.append([row['Date'], row['Day'], row['Holiday']])

wb.save(file_path)
print(f'Updated {file_path}')
print(f'Sheets: {wb.sheetnames}')
for name, df in sheets_to_add.items():
    print(f'  {name}: {len(df)} holidays')
